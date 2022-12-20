from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/search")
def search():
    query = request.args.get("q")
    gender = request.args.get("gender")

    if not query:
        return "", 400

    names = find(query, gender)
    print(names)
    for x in range(10):
        print("")

    return jsonify(names)


def find(query, gender):
    names = list()
    query = query.capitalize()
    wb = load_workbook("data/names.xlsx")
    print(wb.sheetnames)
    for x in range(10):
        print("")

    if gender == "1" or gender == "0":
        wsG = wb.active
        colB = wsG["B"]
        colC = wsG["C"]

        for name, disc in zip(colB, colC):
            if name.value.startswith(query):
                names.append([name.value, disc.value])

    if gender == "2" or gender == "0":
        wsB = wb["Boys"]
        colB = wsB["B"]
        colC = wsB["C"]

        for name, disc in zip(colB, colC):
            print(name.value)
            if name.value.startswith(query):
                names.append([name.value, disc.value])


    return names









